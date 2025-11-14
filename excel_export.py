"""
Optional Excel export functionality
This module provides functionality to export messages to Excel files.
"""

import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

def create_excel_file(filename=None):
    """Create a new Excel file with headers."""
    if filename is None:
        filename = f"fb_messages_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Messages"
    
    # Add headers matching Google Sheets format
    headers = ['Timestamp', 'Name', 'Contact Number', 'Purok', 'Barangay', 'Mensahe', 'Please Upload Picture', 'Email Address', 'Status']
    ws.append(headers)
    
    # Style headers
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    wb.save(filename)
    return filename, wb, ws

def append_to_excel(filename, message_data):
    """Append message data to an Excel file with parsed structured data."""
    try:
        # Import message parser
        from message_parser import parse_message_data
        
        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            # Check if headers match
            try:
                first_row = [cell.value for cell in ws[1]]
                expected_headers = ['Timestamp', 'Name', 'Contact Number', 'Purok', 'Barangay', 'Mensahe', 'Please Upload Picture', 'Email Address', 'Status']
                if first_row != expected_headers:
                    # Clear and add correct headers
                    ws.delete_rows(1, ws.max_row)
                    ws.append(expected_headers)
                    # Style headers
                    from openpyxl.styles import Font, PatternFill
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
            except:
                pass
        else:
            filename, wb, ws = create_excel_file(filename)
        
        # Extract basic message information
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sender_name = message_data.get('from', {}).get('name', 'Unknown')
        message_text = message_data.get('message', {}).get('text', '')
        attachments = message_data.get('message', {}).get('attachments', [])
        
        # Parse message to extract structured data
        parsed_data = parse_message_data(message_text, sender_name, attachments)
        
        # Append row with new format
        row = [
            timestamp,                          # Timestamp
            parsed_data['name'],                # Name
            parsed_data['contact_number'],      # Contact Number
            parsed_data['purok'],               # Purok
            parsed_data['barangay'],            # Barangay
            parsed_data['mensahe'],             # Mensahe (Message)
            parsed_data['attachment_url'],      # Please Upload Picture (Attachment URL)
            parsed_data['email'],               # Email Address
            parsed_data['status']               # Status
        ]
        ws.append(row)
        
        # Save
        wb.save(filename)
        return True
    except Exception as e:
        print(f"Error appending to Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

